'''
Main module of package. This module constructs, evaluates, exports, and utilises
NLP programs.

Classes:

    NLP

Functions:

    startRec() -> float
    stopRec(float) -> tuple[float, float]

Misc variables:

    cwd : str
        A string containing the path to the current working directory when the
        module is first run.
    all_possible_params : list
        The names of all possible reportable input parameters.
    space_measures : list
        The names of all possible peak space use measures.
    all_possible_measures : list
        The names of all possible performance measures.

Exceptions:

    ImporterException
    VectoriseException
    MLearnException
    CrossValidateException
    NameExistsException
'''
import os

import openpyxl.styles
from .Importer import importer as i
from .vectorise import vectorise as v
from .mlearn import mlearn as m
from .evaluate import evaluate as e
from .evaluate import base as eb
from . import exceptions as ex
import joblib
import time
import tracemalloc
import openpyxl
import pandas as pd
import csv
from scipy import sparse
from sklearn.model_selection import KFold

class NLP:
    '''
    A class to represent an NLP program.

    ...

    Attributes
    ----------
    name : str
        The name of the NLP program.
    parameters : dict
        A dictionary containing the parameters used to build the NLP program.
    importer : Importer
        The importer used to extract the data.
    vectorise : Vectorise
        The vectoriser used to vectorise the data.
    mlearn : MLearn
        The machine learning model used to predict flags.

    Constructed by running the create() method:
    trainData : list
        A list of records used to train the model and evaluate the complexity.
    vectTime : float
        The time taken to vectorise the training data.
    vectSpace : float
        The peak memory used when vectorising the training data.
    trainedModel : Union[DecisionTreeClassifier,
                        BalancedRandomForestClassifier,
                        RUSBoostClassifier,
                        SVC]
        The useable machine learning model which can calssify input vectors.
    actualFlags : list
        The flags from the training data.
    predictedFlags : list
        The predicted flags for the training data produced by the NLP 
        program.
    times : list
        The measured times taken by different processes.
    spaces : list
        The measured peak memory usages of different processes.

    Constructed by running the crossValidate() method:
    nFolds : int
        The number of folds of training data to use for cross validation.
    trainData : list
        A list of records used to train the model and evaluate the complexity.
    vectTime : float
        The time taken to vectorise the training data.
    vectSpace : float
        The peak memory used when vectorising the training data.

    Constructed by running the evaluateNLP() method:
    eval : Evaluate
        The evaluater used to evaluate the program.
    outputDict : dict
        A dictionary containing the evaluation measures of the NLP program.
    
    Constructed by running the exportNLP() method.
    cwd : str
        The directory from which base is run.
    directoryName : str
        The name of the directory where files are exported to.

    Methods
    -------
    create()
        Imports, vectorises and uses data to train ML component. Then makes 
        predictions and constructs attributes.
    crossValidate(int) -> dict
        Imports, vectorises and uses data to train and evaluate an NLP program
        through cross validation.
    evaluateNLP()
        Evaluates the NLP program and constructs final attributes.
    exportNLP()
        Pickles and exports the NLP object to the model directory.
    annotateDataCSV(str, list, str)
        Write classifications into a CSV file of records.
    annotateDataXLSX(str, str, list, str)
        Write classifications into a XLSX file of records.
    crossValidateQuantify(int) -> dict
        Compares actual and estimated positive instance counts across multiple folds.
    viewEvaluation()
        Prints the evaluation to the command line.
    viewParameters()
        Prints the parameters to the command line.
    startExcel(str)
        Creates an excel file for recording program parameters and performance.
    writeToExcel(str)
        Writes program parameters and evaluation into an excel file.
    '''
    def __init__(self, 
                 name : str = 'program',
                 imp_arg_dict : dict = None,
                 trainFile : str = '',
                 testFile : str = '',
                 fileLocations: list = [],
                 dateColumnLabel : str = '',
                 hospitalColumnLabel : str = '',
                 sexColumnLabel : str = '',
                 ageColumnLabel : str = '',
                 customColumnLabels : list = [],
                 textFieldColumnLabels : list = [],
                 flagColumnLabel : str = '', 
                 ageBounds : tuple = (0, 150), 
                 hospital: str = 'ALL', 
                 sex: int = 0, 
                 yearBounds : tuple = (1900, 2100),
                 customBounds : list = [], 
                 trainSize: int = 0, 
                 trainDist: str = 'NEWESTBLOCK', 
                 testSize: int = 0, 
                 testDist: str = 'NEWESTBLOCK',
                 vect_arg_dict : dict = None,
                 tokeniser : str = '',
                 preLAChanges : list = [],
                 tokenLevelLA : list = [],
                 textLevelLA : list = [],
                 corpusLevelLA : str = '',
                 ngramRange : tuple = (1, 1),
                 ml_arg_dict : dict = None,
                 mlAlgType : str = '',
                 macLearnInput : dict = {},
                 overSampleOps : dict = {},
                 underSampleOps : dict = {}):
        '''
        Checks name input valid and contructs initial attributes for NLP object.

        Parameters
        ----------
        name : str
            A name for the NLP program.
        imp_arg_dict : dict
            A dictionary containing Importer constructor arguments.
        trainFile : str
            The path to a file containing training data.
        testFile : str
            The path to a file containing testing data.
        fileLocations : list
            List of paths of files where the data is stored.
        dateColumnLabel : str
            The label of the column with the date data.
        hospitalColumnLabel : str
            The label of the column with the hospital data.
        sexColumnLabel : str
            The label of the column with the sex data.
        ageColumnLabel : str
            The label of the column with the age data.
        customColumnLabels : list
            The labels of columns with custom data to filter.
        textFieldColumnLabel : list
            The labels of the columns with free text.
        flagColumnLabel : str
            The label of the column with the classification flag.
        ageBounds : tuple
            The minimum and maximum age values for included records.
        hospital: str
            Which hospital(s) to include.
        sex: int
            Which sex(es) to include.
        yearBounds : tuple
            The minimum and maximum year values for included records.
        customBounds : list
            A list of bounds for columns with custom data.
        trainSize: int
            The amount of training records. 
        trainDist: str
            The distribution to use for selecting the training records. 
        testSize: int
            The amount of testing records.
        testDist: str
            The distribution to use for selecting testing records.
        vect_arg_dict : dict
            A dictionary containing Vectorise constructor arguments.
        tokeniser : str
            A choice of tokeniser to use for vectorising.
        preLAChanges : list
            A choice of pre LA changes for vectorising.
        tokenLevelLA : list
            A choice of token level techniques for vectorising.
        textLevelLA : list
            A choice of text level techniques for vectorising.
        corpusLevelLA : str
            A choice of corpus level technique for vectorising.
        ngramRange : tuple
            Lower and upper bound for n-gram sizes (inclusive).
        ml_arg_dict : dict
            A dictionary containing MLearn constructor arguments.
        mlAlgType : str
            A choice of machine learning algorithm type.
        macLearnInput : dict
            Additional machine learning algorithm specifications.
        overSampleOps : dict
            A dictionary of input parameters relavent to over-sampling.
        underSampleOps : dict
            A dictionary of input parameters relavent to under-sampling.
        
        NOTE : If an argument dictionary is provided for Importer, Vectorise, or MLearn,
        the related keyword arguments provided to the NLP constructor will be ignored.
        '''
        self.name = name

        self.parameters = {'Name' : self.name}

        if imp_arg_dict != None:
            if not isinstance(imp_arg_dict, dict):
                raise ex.ImporterException()
            self.importer = i.Importer(imp_arg_dict)
            for key, val in imp_arg_dict.items():
                self.parameters[key] = val
        else:
            self.importer = i.Importer(trainFile= trainFile,
                                    testFile= testFile,
                                    fileLocations= fileLocations, 
                                    dateColumnLabel= dateColumnLabel,
                                    hospitalColumnLabel= hospitalColumnLabel,
                                    sexColumnLabel= sexColumnLabel,
                                    ageColumnLabel= ageColumnLabel,
                                    customColumnLabels= customColumnLabels,
                                    textFieldColumnLabels= textFieldColumnLabels,
                                    flagColumnLabel= flagColumnLabel,
                                    ageBounds= ageBounds,
                                    hospital= hospital,
                                    sex= sex,
                                    yearBounds= yearBounds,
                                    customBounds= customBounds,
                                    trainSize= trainSize,
                                    trainDist= trainDist,
                                    testSize= testSize,
                                    testDist= testDist
                                    )
            imp_params = dict(
                trainFile = trainFile,
                testFile = testFile,
                fileLocations = fileLocations,
                dateColumnLabel = dateColumnLabel,
                hospitalColumnLabel = hospitalColumnLabel,
                sexColumnLabel = sexColumnLabel,
                ageColumnLabel = ageColumnLabel,
                textFieldColumnLables = textFieldColumnLabels,
                flagColumnLabel = flagColumnLabel,
                customColumnLabels = customColumnLabels,
                ageBounds = ageBounds,
                hospital = hospital,
                sex = sex,
                yearBounds = yearBounds,
                customBounds = customBounds,
                trainSize = trainSize,
                trainDist = trainDist,
                testSize = testSize,
                testDist = testDist
            )
            self.parameters.update(imp_params)

        if vect_arg_dict != None:
            if not isinstance(vect_arg_dict, dict):
                raise ex.VectoriseException()
            self.vectorise = v.Vectorise(vect_arg_dict)
            for key, val in vect_arg_dict.items():
                self.parameters[key] = val
        else:
            self.vectorise = v.Vectorise(tokeniser= tokeniser,
                                         preLAChanges= preLAChanges, 
                                         tokenLevelLA= tokenLevelLA, 
                                         textLevelLA= textLevelLA,
                                         corpusLevelLA= corpusLevelLA,
                                         ngramRange= ngramRange)
            vect_params = dict(
                tokeniser = tokeniser,
                preLAChanges = preLAChanges,
                tokenLevelLA = tokenLevelLA,
                textLevelLA = textLevelLA,
                corpusLevelLA = corpusLevelLA,
                ngramRange = ngramRange
            )
            self.parameters.update(vect_params)

        if ml_arg_dict != None:
            if not isinstance(ml_arg_dict, dict):
                raise ex.MLearnException()
            self.mlearn = m.MLearn(arg_dict= ml_arg_dict)
            for key, val in ml_arg_dict.items():
                self.parameters[key] = val
        else:
            self.mlearn = m.MLearn(mlAlgType= mlAlgType, 
                                   macLearnInput= macLearnInput, 
                                   overSampleOps= overSampleOps, 
                                   underSampleOps= underSampleOps)     
            ml_params = dict(
                mlAlgType = mlAlgType,
                macLearnInput = macLearnInput,
                overSampleOps = overSampleOps,
                underSampleOps = underSampleOps
            )
            self.parameters.update(ml_params)

    def create(self):
        '''
        Imports, vectorises and uses data to train ML component. Then makes 
        predictions and constructs attributes.

        Parameters
        ----------
        None

        Returns
        -------
        None
        '''
        self.trainData, testData = self.importer.importData()

        trainFlags = self.trainData[self.importer.flagColumnLabel].values.tolist()
        actualFlags = testData[self.importer.flagColumnLabel].values.tolist()

        self.trainData = self.trainData[self.importer.textFieldColumnLabels].values.tolist()
        testData = testData[self.importer.textFieldColumnLabels].values.tolist()

        print('Vectorising Data...')
        time0 = startRec()
        trainVectors, testVectors = self.vectorise.vectorise(self.trainData, 
                                                             testData)
        self.vectTime, self.vectSpace = stopRec(time0)

        predictedFlags, trainedModel = self.mlearn.trainAndPredict(trainVectors, 
                                                                trainFlags, 
                                                                testVectors)

        times = [self.importer.importTime, 
                 self.importer.filterTime,
                 self.importer.trainExtractTime,
                 self.importer.testExtractTime,
                 self.vectTime,
                 self.mlearn.trainingTime,
                 self.mlearn.predictionTime]
        
        spaces = [self.importer.importSpace,
                  self.importer.filterSpace,
                  self.importer.trainExtractSpace,
                  self.importer.testExtractSpace,
                  self.vectSpace,
                  self.mlearn.trainingSpace,
                  self.mlearn.predictionSpace]

        self.trainedModel = trainedModel
        self.actualFlags = actualFlags
        self.mlearn.testFlags = actualFlags
        self.predictedFlags = predictedFlags
        self.times = times
        self.spaces = spaces
        print('NLP Program Created')

    def crossValidate(self, 
                      nFolds : int= 5) -> dict:
        '''
        Imports, vectorises and uses data to train and evaluate an NLP program
        through cross validation.
        
        Parameters
        ----------
        nFolds : int
            The number of folds of training data to use for cross validation.

        Returns
        -------
        scores : dict
            A dictionary containing measures evaluated during cross validation.
        '''
        if not isinstance(nFolds, int):
            raise ex.CrossValidateException(
                'nFolds must be an int.'
            )
        if nFolds < 1:
            raise ex.CrossValidateException(
                'nFolds must be greater than or equal to 1.'
            )
        
        self.nFolds = nFolds
        self.trainData, testData = self.importer.importData()

        trainFlags = self.trainData[self.importer.flagColumnLabel].values.tolist()

        self.trainData = self.trainData[self.importer.textFieldColumnLabels].values.tolist()
        testData = testData[self.importer.textFieldColumnLabels].values.tolist()

        print('Vectorising Data...')
        time0 = startRec()
        trainVectors, _ = self.vectorise.vectorise(self.trainData, 
                                                    testData)
        self.vectTime, self.vectSpace = stopRec(time0)

        scores = {'Train F1' : [],
                  'Train Precision' : [],
                  'Train Recall' : [],
                  'Test F1' : [],
                  'Test Precision' : [],
                  'Test Recall' : [],
                  'Fold Times' : []}

        print('Cross-Validating...')
        kf = KFold(n_splits= nFolds)
        for i, (trainIndex, testIndex) in enumerate(kf.split(trainFlags)):
            print(f'Running on Fold {i + 1}...')

            time0 = time.time()
            
            subTrainVectors = sparse.vstack([trainVectors.getrow(i) for i in trainIndex])
            subTrainFlags = [trainFlags[i] for i in trainIndex]
            subTestVectors = sparse.vstack([trainVectors.getrow(i) for i in testIndex])
            subTestFlags = [trainFlags[i] for i in testIndex]

            predictedTestFlags, _ = self.mlearn.trainAndPredict(subTrainVectors, 
                                                            subTrainFlags, 
                                                            subTestVectors)
            
            print('Predicting on Training Vectors...')
            predictedTrainFlags = self.mlearn.trainedModel.predict(subTrainVectors)

            trainPrecision, trainRecall = eb.precisionAndRecall(subTrainFlags,
                                                                predictedTrainFlags)
            
            if trainPrecision == 0 and trainRecall == 0:
                trainf1 = 0
            else:
                trainf1 = (2 * trainPrecision * trainRecall) / (trainPrecision + trainRecall)

            scores['Train F1'].append(trainf1)
            scores['Train Precision'].append(trainPrecision)
            scores['Train Recall'].append(trainRecall)
            
            testPrecision, testRecall = eb.precisionAndRecall(subTestFlags,
                                                              predictedTestFlags)
            
            if testPrecision == 0 and testRecall == 0:
                testf1 = 0
            else:
                testf1 = (2 * testPrecision * testRecall) / (testPrecision + testRecall)
        
            scores['Test F1'].append(testf1)
            scores['Test Precision'].append(testPrecision)
            scores['Test Recall'].append(testRecall)
            scores['Fold Times'].append(time.time() - time0)

        print('NLP Program Cross-Validated')
        if hasattr(self, 'outputDic'):
            self.outputDic.update(scores)
        else:
            self.outputDic = scores
        return scores
    
    def evaluateNLP(self):
        '''
        Evaluates the NLP program and constructs additional attributes.

        Parameters
        ----------
        None

        Returns
        -------
        None
        '''
        print('Evaluating Model...')
        time0 = startRec()
        self.eval = e.Evaluate(self.actualFlags, 
                                self.predictedFlags,
                                self.times,
                                self.spaces)
        self.outputDic = self.eval.evaluate()
        evalTime, evalSpace = stopRec(time0)
        self.eval.times.append(evalTime)
        self.eval.spaces.append(evalSpace)
        self.outputDic['EvaluateTime'] = evalTime
        self.outputDic['EvaluateSpace'] = evalSpace

    def exportNLP(self):
        '''
        Pickles and exports the NLP object to the model directory.

        Parameters
        ----------
        None

        Returns
        -------
        None
        '''
        self.cwd = cwd
        self.directoryName = self.name + '-NLP-Program'
        path = str(os.getcwd()) + '\\EpiNLPpb_dev\\model'
        os.chdir(path)

        if os.path.exists(str(os.getcwd()) + '\\' + self.directoryName):
            raise ex.NameExistsException(
                'There is already an NLP program stored with this name.'
                )

        os.makedirs(self.directoryName)
        path = (
            str(os.getcwd()) + '\\' + self.directoryName
                )
        os.chdir(path)
        print('Exporting NLP program...')
        joblib.dump(self, "nlp.pkl")
        os.chdir(cwd)
        print('Done')

    def annotateDataCSV(self, 
                        path : str, 
                        textFieldColumnLabels : list,
                        columnLabel : str):
        '''
        Write classifications into a CSV file of records.

        Parameters
        ----------
        path : str
            The location of the CSV file to be annotated.
        textFieldColumnLabels : list
            The lables of the columns containg the text to be used.
        columnLabel : str
            A label for the column of annotations.

        Returns
        -------
        None
        '''
        print('Importing text...')
        outList = pd.read_csv(
                    path, 
                    usecols = textFieldColumnLabels, 
                    encoding_errors= 'ignore', 
                    low_memory= False
                    ).values.tolist()
        print('Vectorising...')
        vectors = self.vectorise.vectoriseList(outList)
        print('Predicting...')
        flags = self.trainedModel.predict(vectors)
        print('Annotating file...')
        csv_file_path = path
        with open(file= csv_file_path, mode= 'r', errors= 'replace') as file:
            reader = csv.reader(file)
            data = list(reader)
        data[0].append(columnLabel)
        rowIndex = 1
        for flag in flags:
            data[rowIndex].append(flag)
            rowIndex += 1
        print('Saving file...')
        with open(csv_file_path, 'w', newline='', errors= 'replace') as file:
            writer = csv.writer(file)
            writer.writerows(data)
        print('Data annotated.')

    def annotateDataXLSX(self, 
                        path : str, 
                        sheet : str, 
                        textFieldColumnLabels : list,
                        columnLabel : str):
        '''
        Write classifications into an XLSX file of records.

        Parameters
        ----------
        path : str
            The location of the XLSX file to be annotated.
        sheet : str
            The name of the sheet to be annotated.
        textFieldColumnLabels : list
            The lables of the columns containg the text to be used.
        columnLabel : str
            A label for the column of annotations.

        Returns
        -------
        None
        '''
        print('Opening file...')
        file = openpyxl.load_workbook(path)
        sheet_name = sheet
        sheet = file[sheet]
        print('Importing text...')
        outList = pd.read_excel(
                    path,
                    sheet_name= sheet_name, 
                    usecols = textFieldColumnLabels
                    ).values.tolist()
        print('Vectorising...')
        vectors = self.vectorise.vectoriseList(outList)
        print('Predicting...')
        flags = self.trainedModel.predict(vectors)
        print('Annotating file...')
        colIndex = 1
        while sheet.cell(row = 1, column = colIndex).value != None:
            colIndex += 1
        sheet.cell(row = 1, column = colIndex).value = columnLabel
        rowIndex = 2
        for flag in flags:
            sheet.cell(row = rowIndex, column = colIndex).value = flag
            rowIndex += 1
        print('Saving file...')
        file.save(path)

    def crossValidateQuantify(self,
                              nFolds : int = 5) -> dict:
        '''
        Compares actual and estimated positive instance counts across multiple folds.

        Parameters
        ----------
        nFolds : int
            The number of folds of training data to use for cross validation.

        Returns
        counts : dict
            A dictionary containing actual and estimated counts for each fold.
        '''
        if not isinstance(nFolds, int):
            raise ex.CrossValidateException(
                'nFolds must be an int.'
            )
        if nFolds < 1:
            raise ex.CrossValidateException(
                'nFolds must be greater than or equal to 1.'
            )
        
        self.trainData, testData = self.importer.importData()

        trainFlags = self.trainData[self.importer.flagColumnLabel].values.tolist()

        self.trainData = self.trainData[self.importer.textFieldColumnLabels].values.tolist()
        testData = testData[self.importer.textFieldColumnLabels].values.tolist()

        print('Vectorising Data...')
        time0 = startRec()
        trainVectors, _ = self.vectorise.vectorise(self.trainData, 
                                                    testData)
        self.vectTime, self.vectSpace = stopRec(time0)

        counts = {'Actual Counts' : [],
                  'Estimated Counts' : []}

        print('Cross-Validating...')
        kf = KFold(n_splits= nFolds)
        for i, (trainIndex, testIndex) in enumerate(kf.split(trainFlags)):
            print(f'Running on Fold {i + 1}...')

            time0 = time.time()
            
            subTrainVectors = sparse.vstack([trainVectors.getrow(i) for i in trainIndex])
            subTrainFlags = [trainFlags[i] for i in trainIndex]
            subTestVectors = sparse.vstack([trainVectors.getrow(i) for i in testIndex])
            subTestFlags = [trainFlags[i] for i in testIndex]

            predictedTestFlags, _ = self.mlearn.trainAndPredict(subTrainVectors, 
                                                            subTrainFlags, 
                                                            subTestVectors)
            
            counts['Actual Counts'].append(sum(subTestFlags))
            counts['Estimated Counts'].append(sum(predictedTestFlags))

            testPrecision, testRecall = eb.precisionAndRecall(subTestFlags,
                                                              predictedTestFlags)
            
            if testPrecision == 0 and testRecall == 0:
                testf1 = 0
            else:
                testf1 = (2 * testPrecision * testRecall) / (testPrecision + testRecall)

            print('Fold F1-score: ', testf1)

        return counts
    
    def viewEvaluation(self):
        '''
        Prints the evaluation to the command line.

        Parameters
        ----------
        None

        Returns
        -------
        None
        '''
        print('Precision: ', round(self.outputDic['Precision']*100, 2), ' %')
        print('Recall: ', round(self.outputDic['Recall']*100, 2), ' %')
        print('Cross Validation:')
        if not hasattr(self, 'nFolds'):
            self.nFolds = 0
        for i in range(0, self.nFolds):
            print(f' Fold {i}:')
            print(' Train Precision: ', self.outputDic['Train Precision'][i])
            print(' Trian Recall: ', self.outputDic['Train Recall'][i])
            print(' Test Precision: ', self.outputDic['Test Precision'][i])
            print(' Test Recall: ', self.outputDic['Test Recall'][i])
            print(' Time: ', self.outputDic['Fold Times'][i], '(s)')
        print('Total', '\n', 'Time: ', 
            round(self.outputDic['TotalTime'],3), '(s) Space: ', 
            round(self.outputDic['TotalSpace'] / (1024*1024),3), 
            '(MB)')
        print('Importing', '\n', 'Time: ', 
            round(self.outputDic['ImportTime'],3), '(s) Space: ', 
            round(self.outputDic['ImportSpace'] / (1024*1024),3), 
            '(MB)')
        print('Filtering', '\n', 'Time: ', 
            round(self.outputDic['FilterTime'],3), '(s) Space: ', 
            round(self.outputDic['FilterSpace'] / (1024*1024),3), 
            '(MB)')
        print('Extracting Training Data ', '\n', 'Time: ', 
            round(self.outputDic['TrainExtractTime'],3), '(s) Space: ', 
            round(self.outputDic['TrainExtractSpace'] / (1024*1024),3), 
            '(MB)')
        print('Extracting Testing Data ', '\n', 'Time: ', 
            round(self.outputDic['TestExtractTime'],3), '(s) Space: ', 
            round(self.outputDic['TestExtractSpace'] / (1024*1024),3), 
            '(MB)')
        print('Data Vectorising ', '\n', 
            round(self.outputDic['VectoriseTime'],3), '(s) Space: ', 
            round(self.outputDic['VectoriseSpace'] / (1024*1024),3), 
            '(MB)')
        print('ML Algorithm Training', '\n', 'Time: ', 
            round(self.outputDic['MLTrainingTime'],3), '(s) Space: ', 
            round(self.outputDic['MLTrainingSpace'] / (1024*1024),3), 
            '(MB)')
        print('ML Algorithm Predicting', '\n', 'Time: ', 
            round(self.outputDic['MLPredictionTime'],3), '(s) Space: ', 
            round(self.outputDic['MLPredictionSpace'] / (1024*1024),3), 
            '(MB)')
        print('Evaluating', '\n', 'Time: ', 
            round(self.outputDic['EvaluateTime'],3), '(s) Space: ', 
            round(self.outputDic['EvaluateSpace'] / (1024*1024),3), 
            '(MB)')
        print('----------------------------------------------------------------'
              '---------------------------------------------------------')
    
    def viewParameters(self):
        '''
        Prints the parameters to the command line.

        Parameters
        ----------
        None

        Returns
        -------
        None
        '''
        for key, val in self.parameters.items():
            print(f'{key} : {val}')
        print('----------------------------------------------------------------'
              '---------------------------------------------------------')

    def startExcel(self,
                   file : str):
        '''
        Creates an excel file for recording program parameters and performance.

        Parameters
        ----------
        file : str
            The path at which to create the excel file.

        Returns
        -------
        None
        '''
        wb = openpyxl.Workbook()
        ws = wb.active

        row = all_possible_params + all_possible_measures
        ws.append(row)

        for id, _ in enumerate(row):
            cell = ws.cell(row= 1, column= id + 1)
            cell.font = openpyxl.styles.Font(bold= True)

        wb.save(file)
    
    def writeToExcel(self,
                     file : str):
        '''
        Writes program parameters and evaluation into model/file.

        Parameters
        ----------
        file : str
            The path of the excel file to write results to.

        Returns
        -------
        None
        '''
        row = []
        for param in all_possible_params:
            if param in self.parameters:
                row.append(str(self.parameters[param]))
            elif param in self.parameters['macLearnInput']:
                row.append(str(self.parameters['macLearnInput'][param]))
            else:
                row.append('')

        for space in space_measures:
            if space in self.outputDic:
                self.outputDic[space] = round(self.outputDic[space] / (1024*1024), 3)
        
        for measure in all_possible_measures:
            if measure in self.outputDic:
                row.append(str(self.outputDic[measure]))
            else:
                row.append('')            

        wb = openpyxl.load_workbook(file)
        ws = wb.active
        ws.append(row)

        wb.save(file)

def startRec() -> float:
    '''
    Outputs the current time and begins tracking memory.

    Parameters
    ----------
    None

    Retruns
    -------
    time0 : float
        The current time.
    '''
    time0 = time.time()
    tracemalloc.start()
    return time0

def stopRec(time0 : float) -> tuple[float, float]:
    '''
    Outputs time passed and peak memory used since the last call of startRec().

    Parameters
    ----------
    time0 : float
        The output from the last call of startRec().

    Returns
    -------
    time1 : float
        The time passed since the last call of startRec().
    space : float
        The peak memory used since the last call of startRec().
    '''
    time1 = time.time() - time0
    _, space = tracemalloc.get_traced_memory()
    tracemalloc.stop()
    return (time1, space)

cwd = os.getcwd()

all_possible_params = [
    'trainFile',
    'testFile',
    'fileLocations', 
    'dateColumnLabel',
    'hospitalColumnLabel',
    'sexColumnLabel',
    'ageColumnLabel',
    'customColumnLabels',
    'textFieldColumnLabels',
    'flagColumnLabel',
    'ageBounds',
    'hospital',
    'sex',
    'yearBounds',
    'customBounds',
    'trainSize',
    'trainDist',
    'testSize',
    'testDist',
    'tokeniser',
    'preLAChanges', 
    'tokenLevelLA', 
    'textLevelLA',
    'corpusLevelLA',
    'ngramRange',
    'mlAlgType',
    'impurity',
    'ratio',
    'max_depth',
    'min_samples_split',
    'n_estimators',
    'learning_rate'
    'class_weight',
    'kernel',
    'gamma',
    'degree',
    'r',
    'C',
    'overSampleOps',
    'underSampleOps'        
]

space_measures = [
    'TotalSpace',
    'ImportSpace',
    'FilterSpace',
    'ExtractSpace',
    'VectoriseSpace',
    'MLTrainingSpace',
    'MLPredictionSpace',
    'EvaluateSpace'
]

all_possible_measures = [
    'Precision',
    'Recall',
    'TotalTime',
    'TotalSpace',
    'Train F1',
    'Train Precision',
    'Train Recall',
    'Test F1',
    'Test Precision',
    'Test Recall',
    'Fold Times',
    'ImportTime',
    'ImportSpace',
    'FilterTime',
    'FilterSpace',
    'ExtractTime',
    'ExtractSpace',
    'VectoriseTime',
    'VectoriseSpace',
    'MLTrainingTime',
    'MLTrainingSpace',
    'MLPredicitonTime',
    'MLPredictionSpace',
    'EvaluateTime',
    'EvaluateSpace'
]